"""
Serviço responsável por exportar o ranking de fundos para Excel (.xlsx).

Mantido separado de fundos_service.py porque lida com uma preocupação
totalmente diferente (formatação de planilha), não com cálculo financeiro.

Como calcular a volatilidade de todos os fundos pode demorar (histórico
diário de cada um), a exportação roda em background numa thread e expõe um
sistema simples de "jobs" em memória para o front-end acompanhar o
progresso via polling.
"""
import threading
import uuid
from io import BytesIO
from typing import List, Dict, Any, Optional, Callable

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from services.fundos_service import calcular_volatilidade_ranking_fundo

# Mesmos 5 períodos usados no ranking (fundos_service.PESOS_RANKING_PADRAO)
PERIODOS_EXPORTACAO = ["12m", "24m", "36m", "48m", "60m"]

ROTULOS_PERIODOS = {
    "12m": "12 meses",
    "24m": "24 meses",
    "36m": "36 meses",
    "48m": "48 meses",
    "60m": "60 meses",
}


def _obter_volatilidades_fundo(
    fundo: Dict[str, Any],
    data_referencia: Optional[str],
) -> Dict[str, Any]:
    """
    Retorna as volatilidades dos 5 períodos para um fundo do ranking.

    Se o front-end já enviou os valores (usuário clicou em "Ver
    Volatilidade" antes de exportar), reaproveita — evita recalcular à toa,
    já que isso exige carregar o histórico diário do fundo. Caso contrário,
    calcula sob demanda aqui mesmo.
    """
    chaves = [f"volatilidade_{periodo}" for periodo in PERIODOS_EXPORTACAO]

    if all(fundo.get(chave) is not None for chave in chaves):
        return {chave: fundo[chave] for chave in chaves}

    resultado = calcular_volatilidade_ranking_fundo(
        cnpj=fundo["cnpj"],
        data_referencia=data_referencia,
    )
    return {chave: resultado.get(chave) for chave in chaves}


def gerar_excel_ranking(
    fundos: List[Dict[str, Any]],
    data_referencia: Optional[str] = None,
    categoria: Optional[str] = None,
    progresso_callback: Optional[Callable[[int], None]] = None,
) -> BytesIO:
    """
    Gera um arquivo .xlsx em memória com o ranking de fundos: CNPJ, nome,
    tipo (classificação ANBIMA) e rentabilidade + volatilidade dos cinco
    períodos padrão (12/24/36/48/60 meses), além da nota final.

    `fundos` é a lista de dicionários no mesmo formato retornado por
    /api/fundos/ranking, podendo já conter volatilidade_XXm (se o usuário
    já tiver consultado) ou não.

    Se `progresso_callback` for informado, é chamado após cada fundo
    processado com o total de fundos já concluídos até o momento — usado
    para reportar o progresso da exportação ao front-end.
    """
    if not fundos:
        raise ValueError("Nenhum fundo informado para exportação")

    wb = Workbook()
    ws = wb.active
    ws.title = "Ranking de Fundos"

    cabecalho = ["#", "CNPJ", "Nome do Fundo", "Tipo"]
    # Bloco 1: todas as rentabilidades (12m, 24m, 36m, 48m, 60m)
    for periodo in PERIODOS_EXPORTACAO:
        rotulo = ROTULOS_PERIODOS[periodo]
        cabecalho.append(f"Rentabilidade {rotulo} (%)")
    # Bloco 2: todas as volatilidades, na mesma ordem de períodos
    for periodo in PERIODOS_EXPORTACAO:
        rotulo = ROTULOS_PERIODOS[periodo]
        cabecalho.append(f"Volatilidade {rotulo} (%)")
    cabecalho.append("Nota Final")

    ws.append(cabecalho)

    # Estilo do cabeçalho
    fonte_cabecalho = Font(bold=True, color="FFFFFF")
    preenchimento_cabecalho = PatternFill(start_color="121212", end_color="121212", fill_type="solid")
    for celula in ws[1]:
        celula.font = fonte_cabecalho
        celula.fill = preenchimento_cabecalho
        celula.alignment = Alignment(horizontal="center", vertical="center")

    for indice, fundo in enumerate(fundos, start=1):
        volatilidades = _obter_volatilidades_fundo(fundo, data_referencia)

        linha = [
            indice,
            fundo.get("cnpj", ""),
            fundo.get("nome", ""),
            fundo.get("tipo") or categoria or "",
        ]

        # IMPORTANTE: a ordem aqui precisa bater com a ordem do cabeçalho
        # acima — dois blocos separados (todas as rentabilidades primeiro,
        # depois todas as volatilidades), e não intercalado por período.
        for periodo in PERIODOS_EXPORTACAO:
            rentabilidade = fundo.get(f"rentabilidade_{periodo}")
            linha.append(rentabilidade if rentabilidade is not None else "")
        for periodo in PERIODOS_EXPORTACAO:
            volatilidade = volatilidades.get(f"volatilidade_{periodo}")
            linha.append(volatilidade if volatilidade is not None else "")

        linha.append(fundo.get("nota_final", ""))
        ws.append(linha)

        if progresso_callback is not None:
            progresso_callback(indice)

    # Formata as colunas numéricas (rentabilidade, volatilidade e nota) com
    # 2 casas decimais. As colunas 5 em diante são as de período; a última
    # coluna é a nota final.
    total_colunas = len(cabecalho)
    for coluna in range(5, total_colunas + 1):
        letra = get_column_letter(coluna)
        for celula in ws[letra][1:]:  # pula o cabeçalho
            if isinstance(celula.value, (int, float)):
                celula.number_format = "0.00"

    # Largura das colunas: #, CNPJ, Nome, Tipo, 5 de rentabilidade,
    # 5 de volatilidade, Nota
    larguras = [5, 20, 40, 16] + [18] * len(PERIODOS_EXPORTACAO) + [18] * len(PERIODOS_EXPORTACAO) + [12]
    for indice, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(indice)].width = largura

    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
# ==============================================================================
# Gerenciamento de jobs de exportação (para reportar progresso ao front-end)
# ==============================================================================

_jobs_lock = threading.Lock()
_jobs: Dict[str, Dict[str, Any]] = {}


def iniciar_exportacao_ranking(
    fundos: List[Dict[str, Any]],
    data_referencia: Optional[str] = None,
    categoria: Optional[str] = None,
) -> str:
    """
    Inicia a geração do Excel em uma thread separada e retorna um job_id
    que o front-end usa para consultar o progresso (obter_status_exportacao)
    e, ao final, baixar o arquivo (obter_arquivo_exportacao).
    """
    if not fundos:
        raise ValueError("Nenhum fundo informado para exportação")

    job_id = uuid.uuid4().hex
    with _jobs_lock:
        _jobs[job_id] = {
            "status": "processando",
            "processados": 0,
            "total": len(fundos),
            "erro": None,
            "buffer": None,
            "nome_arquivo": f"ranking_fundos_{data_referencia or 'atual'}.xlsx",
        }

    thread = threading.Thread(
        target=_executar_exportacao,
        args=(job_id, fundos, data_referencia, categoria),
        daemon=True,
    )
    thread.start()

    return job_id


def _executar_exportacao(
    job_id: str,
    fundos: List[Dict[str, Any]],
    data_referencia: Optional[str],
    categoria: Optional[str],
) -> None:
    """Roda dentro da thread de background; nunca é chamada diretamente."""
    def progresso_callback(processados: int) -> None:
        with _jobs_lock:
            if job_id in _jobs:
                _jobs[job_id]["processados"] = processados

    try:
        buffer = gerar_excel_ranking(
            fundos,
            data_referencia=data_referencia,
            categoria=categoria,
            progresso_callback=progresso_callback,
        )
        with _jobs_lock:
            _jobs[job_id]["status"] = "concluido"
            _jobs[job_id]["buffer"] = buffer
    except Exception as erro:  # noqa: BLE001 - job de background, precisa capturar tudo
        with _jobs_lock:
            _jobs[job_id]["status"] = "erro"
            _jobs[job_id]["erro"] = str(erro)


def obter_status_exportacao(job_id: str) -> Optional[Dict[str, Any]]:
    """Retorna o progresso atual do job, ou None se o job_id não existir."""
    with _jobs_lock:
        job = _jobs.get(job_id)
        if job is None:
            return None
        return {
            "status": job["status"],
            "processados": job["processados"],
            "total": job["total"],
            "erro": job["erro"],
        }


def obter_arquivo_exportacao(job_id: str):
    """
    Retorna (buffer, nome_arquivo) se o job estiver concluído, ou
    (None, None) caso contrário. Remove o job da memória após a entrega,
    já que o buffer só é consumido uma vez.
    """
    with _jobs_lock:
        job = _jobs.get(job_id)
        if job is None or job["status"] != "concluido":
            return None, None
        buffer = job["buffer"]
        nome_arquivo = job["nome_arquivo"]
        del _jobs[job_id]
        return buffer, nome_arquivo