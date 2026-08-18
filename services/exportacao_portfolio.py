import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

# Cores no mesmo padrão visual do resto do app
COR_NEON = "B11116"
COR_HEADER_TEXTO = "FFFFFF"
COR_POSITIVO = "008000"
COR_NEGATIVO = "CC0000"
COR_CINZA_CLARO = "F2F2F2"


def _autofit_colunas(ws, largura_min=10, largura_max=45):
    """Ajusta a largura das colunas com base no conteúdo (aproximado)."""
    for coluna_celulas in ws.columns:
        tamanhos = []
        letra_coluna = None
        for celula in coluna_celulas:
            if letra_coluna is None:
                letra_coluna = getattr(celula, "column_letter", None)
            if letra_coluna is None:
                continue
            valor = celula.value
            if valor is not None:
                tamanhos.append(len(str(valor)))
        if letra_coluna and tamanhos:
            largura = max(largura_min, min(largura_max, max(tamanhos) + 4))
            ws.column_dimensions[letra_coluna].width = largura


def _montar_aba_portfolio(
    wb: Workbook,
    fundos: list[dict],
    data_referencia: str | None,
) -> None:
    ws = wb.active
    ws.title = "Portfólio"

    fonte_padrao = "Arial"

    # Título / metadados
    ws["A1"] = "Composição do Portfólio"
    ws["A1"].font = Font(name=fonte_padrao, size=14, bold=True, color=COR_NEON)

    referencia_texto = data_referencia or datetime.today().strftime("%Y-%m-%d")
    ws["A2"] = f"Data de referência: {referencia_texto}  ·  Gerado em: {datetime.now():%d/%m/%Y %H:%M}"
    ws["A2"].font = Font(name=fonte_padrao, size=9, italic=True, color="666666")

    linha_cabecalho = 4
    colunas = [
        "CNPJ", "Nome do Fundo", "Rentabilidade (36m)", "Volatilidade (36m)",
        "Peso", "Risco Atribuído", "Attribution", "Attribution / Risco",
    ]

    for idx, titulo in enumerate(colunas, start=1):
        celula = ws.cell(row=linha_cabecalho, column=idx, value=titulo)
        celula.font = Font(name=fonte_padrao, size=11, bold=True, color=COR_HEADER_TEXTO)
        celula.fill = PatternFill(start_color=COR_NEON, end_color=COR_NEON, fill_type="solid")
        celula.alignment = Alignment(horizontal="center", vertical="center")

    primeira_linha_dados = linha_cabecalho + 1

    for offset, fundo in enumerate(fundos):
        linha = primeira_linha_dados + offset
        zebra = COR_CINZA_CLARO if offset % 2 == 1 else "FFFFFF"

        ws.cell(row=linha, column=1, value=fundo.get("cnpj", "")).font = Font(name=fonte_padrao, size=10)
        ws.cell(row=linha, column=2, value=fundo.get("nome", "")).font = Font(name=fonte_padrao, size=10)

        rentabilidade = fundo.get("rentabilidade_36m")
        cel_rent = ws.cell(row=linha, column=3)
        if rentabilidade is not None:
            cel_rent.value = round(float(rentabilidade), 4) / 100
            cel_rent.number_format = "0.00%"
            cor = COR_POSITIVO if rentabilidade >= 0 else COR_NEGATIVO
            cel_rent.font = Font(name=fonte_padrao, size=10, bold=True, color=cor)
        else:
            cel_rent.value = "s/ dados"
            cel_rent.font = Font(name=fonte_padrao, size=10, italic=True, color="999999")
        cel_rent.alignment = Alignment(horizontal="center")

        volatilidade = fundo.get("volatilidade_36m")
        cel_vol = ws.cell(row=linha, column=4)
        if volatilidade is not None:
            cel_vol.value = round(float(volatilidade), 4) / 100
            cel_vol.number_format = "0.00%"
        else:
            cel_vol.value = "s/ dados"
            cel_vol.font = Font(name=fonte_padrao, size=10, italic=True, color="999999")
        cel_vol.alignment = Alignment(horizontal="center")
        if volatilidade is not None:
            cel_vol.font = Font(name=fonte_padrao, size=10)

        peso = fundo.get("peso")
        cel_peso = ws.cell(row=linha, column=5)
        if peso is not None:
            cel_peso.value = round(float(peso), 4) / 100
            cel_peso.number_format = "0.00%"
        cel_peso.font = Font(name=fonte_padrao, size=10)
        cel_peso.alignment = Alignment(horizontal="center")

        # Risco Atribuído (soma da coluna do fundo na matriz de
        # covariância ÷ soma total da matriz)
        risco_atribuido = fundo.get("risco_atribuido")
        cel_risco = ws.cell(row=linha, column=6)
        if risco_atribuido is not None:
            cel_risco.value = round(float(risco_atribuido), 4) / 100
            cel_risco.number_format = "0.00%"
            cel_risco.font = Font(name=fonte_padrao, size=10)
        else:
            cel_risco.value = "—"
            cel_risco.font = Font(name=fonte_padrao, size=10, italic=True, color="999999")
        cel_risco.alignment = Alignment(horizontal="center")

        # Attribution = rentabilidade (36m) x peso
        attribution = fundo.get("attribution")
        cel_attribution = ws.cell(row=linha, column=7)
        if attribution is not None:
            cel_attribution.value = round(float(attribution), 4) / 100
            cel_attribution.number_format = "0.00%"
            cor = COR_POSITIVO if attribution >= 0 else COR_NEGATIVO
            cel_attribution.font = Font(name=fonte_padrao, size=10, color=cor)
        else:
            cel_attribution.value = "—"
            cel_attribution.font = Font(name=fonte_padrao, size=10, italic=True, color="999999")
        cel_attribution.alignment = Alignment(horizontal="center")

        # Attribution / |Risk Attribution|
        attribution_por_risco = fundo.get("attribution_por_risco")
        cel_attr_risco = ws.cell(row=linha, column=8)
        if attribution_por_risco is not None:
            cel_attr_risco.value = round(float(attribution_por_risco), 4)
            cel_attr_risco.number_format = "0.0000"
        else:
            cel_attr_risco.value = "—"
            cel_attr_risco.font = Font(name=fonte_padrao, size=10, italic=True, color="999999")
        cel_attr_risco.alignment = Alignment(horizontal="center")
        if attribution_por_risco is not None:
            cel_attr_risco.font = Font(name=fonte_padrao, size=10)

        for col in range(1, 9):
            ws.cell(row=linha, column=col).fill = PatternFill(
                start_color=zebra, end_color=zebra, fill_type="solid"
            )

    ultima_linha = primeira_linha_dados + len(fundos) - 1

    # Total do peso via fórmula (soma real, não valor fixo) — assim continua
    # correto se o usuário editar os pesos direto na planilha depois.
    if fundos:
        linha_total = ultima_linha + 1
        ws.cell(row=linha_total, column=4, value="Soma dos pesos").font = Font(
            name=fonte_padrao, size=10, bold=True
        )
        ws.cell(row=linha_total, column=4).alignment = Alignment(horizontal="right")

        cel_total = ws.cell(
            row=linha_total,
            column=5,
            value=f"=SUM(E{primeira_linha_dados}:E{ultima_linha})",
        )
        cel_total.number_format = "0.00%"
        cel_total.font = Font(name=fonte_padrao, size=10, bold=True)
        cel_total.alignment = Alignment(horizontal="center")

    ws.freeze_panes = f"A{primeira_linha_dados}"
    _autofit_colunas(ws)


def _montar_aba_correlacao(wb: Workbook, fundos: list[dict], correlacao: dict) -> None:
    if not fundos or not correlacao:
        return

    ws = wb.create_sheet("Correlação")
    fonte_padrao = "Arial"

    ws["A1"] = "Matriz de Correlação (retornos diários, 36 meses)"
    ws["A1"].font = Font(name=fonte_padrao, size=14, bold=True, color=COR_NEON)

    cnpjs = [f["cnpj"] for f in fundos]
    nomes = {f["cnpj"]: f.get("nome", f["cnpj"]) for f in fundos}

    linha_cabecalho = 3
    coluna_inicial = 2  # deixa a coluna A para o nome das linhas

    # Cabeçalho (nomes dos fundos nas colunas)
    for idx, cnpj in enumerate(cnpjs):
        celula = ws.cell(row=linha_cabecalho, column=coluna_inicial + idx, value=nomes[cnpj])
        celula.font = Font(name=fonte_padrao, size=9, bold=True, color=COR_HEADER_TEXTO)
        celula.fill = PatternFill(start_color=COR_NEON, end_color=COR_NEON, fill_type="solid")
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    primeira_linha_dados = linha_cabecalho + 1

    for lin_idx, cnpj_linha in enumerate(cnpjs):
        linha = primeira_linha_dados + lin_idx

        cel_nome = ws.cell(row=linha, column=1, value=nomes[cnpj_linha])
        cel_nome.fill = PatternFill(start_color=COR_NEON, end_color=COR_NEON, fill_type="solid")
        cel_nome.font = Font(name=fonte_padrao, size=9, bold=True, color=COR_HEADER_TEXTO)

        for col_idx, cnpj_coluna in enumerate(cnpjs):
            valor = correlacao.get(cnpj_coluna, {}).get(cnpj_linha)
            cel = ws.cell(row=linha, column=coluna_inicial + col_idx)
            if valor is not None:
                cel.value = round(float(valor), 4)
                cel.number_format = "0.00"
            else:
                cel.value = "—"
            cel.alignment = Alignment(horizontal="center")
            cel.font = Font(name=fonte_padrao, size=9)

    ultima_linha = primeira_linha_dados + len(cnpjs) - 1
    ultima_coluna = coluna_inicial + len(cnpjs) - 1

    if len(cnpjs) >= 2:
        intervalo = (
            f"{get_column_letter(coluna_inicial)}{primeira_linha_dados}:"
            f"{get_column_letter(ultima_coluna)}{ultima_linha}"
        )
        # Escala de cor: vermelho (-1) -> branco (0) -> verde (+1),
        # mesmo sentido usado no dashboard (positivo = verde, negativo = vermelho).
        regra = ColorScaleRule(
            start_type="num", start_value=-1, start_color="FF3366",
            mid_type="num", mid_value=0, mid_color="FFFFFF",
            end_type="num", end_value=1, end_color="00E58C",
        )
        ws.conditional_formatting.add(intervalo, regra)

    ws.freeze_panes = ws.cell(row=primeira_linha_dados, column=coluna_inicial).coordinate
    _autofit_colunas(ws, largura_min=10, largura_max=22)


def _montar_aba_covariancia(
    wb: Workbook,
    fundos: list[dict],
    covariancia: dict,
) -> None:
    """
    Mesmo layout da aba de Correlação, mas com a matriz de covariância
    calculada como correlacao(i,j) * volatilidade(i) * volatilidade(j)
    * peso(i) * peso(j) — ou seja, o valor "real" (não normalizado
    entre -1 e 1) de cada par de fundos, já ponderado pelo peso de cada
    um no portfólio.

    Logo abaixo da matriz, adiciona uma linha "Risco Atribuído" com o
    percentual de contribuição de cada fundo para o risco total do
    portfólio (soma da coluna do fundo / soma total da matriz, via
    fórmula SUM do Excel).
    """
    if not fundos or not covariancia:
        return

    ws = wb.create_sheet("Covariância")
    fonte_padrao = "Arial"

    ws["A1"] = "Matriz de Covariância (correlação × volatilidade × peso)"
    ws["A1"].font = Font(name=fonte_padrao, size=14, bold=True, color=COR_NEON)

    linha_cabecalho = 3

    cnpjs = [f["cnpj"] for f in fundos]
    nomes = {f["cnpj"]: f.get("nome", f["cnpj"]) for f in fundos}

    coluna_inicial = 2  # deixa a coluna A para o nome das linhas

    # Cabeçalho (nomes dos fundos nas colunas)
    for idx, cnpj in enumerate(cnpjs):
        celula = ws.cell(row=linha_cabecalho, column=coluna_inicial + idx, value=nomes[cnpj])
        celula.font = Font(name=fonte_padrao, size=9, bold=True, color=COR_HEADER_TEXTO)
        celula.fill = PatternFill(start_color=COR_NEON, end_color=COR_NEON, fill_type="solid")
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    primeira_linha_dados = linha_cabecalho + 1

    for lin_idx, cnpj_linha in enumerate(cnpjs):
        linha = primeira_linha_dados + lin_idx

        cel_nome = ws.cell(row=linha, column=1, value=nomes[cnpj_linha])
        cel_nome.fill = PatternFill(start_color=COR_NEON, end_color=COR_NEON, fill_type="solid")
        cel_nome.font = Font(name=fonte_padrao, size=9, bold=True, color=COR_HEADER_TEXTO)

        for col_idx, cnpj_coluna in enumerate(cnpjs):
            valor = covariancia.get(cnpj_coluna, {}).get(cnpj_linha)
            cel = ws.cell(row=linha, column=coluna_inicial + col_idx)
            if valor is not None:
                cel.value = round(float(valor), 4)
                cel.number_format = "0.00"
            else:
                cel.value = "—"
            cel.alignment = Alignment(horizontal="center")
            cel.font = Font(name=fonte_padrao, size=9)

    ultima_linha_dados = primeira_linha_dados + len(cnpjs) - 1
    ultima_coluna = coluna_inicial + len(cnpjs) - 1

    # ---- Risk Attribution: soma da coluna de cada fundo dividida pela
    # soma total da matriz. Usa fórmulas do Excel (SUM), assim continua
    # correta se o usuário editar algum valor da matriz depois. ----
    if len(cnpjs) >= 2:
        linha_risco = ultima_linha_dados + 2

        cel_titulo = ws.cell(row=linha_risco, column=1, value="Risco Atribuído")
        cel_titulo.font = Font(name=fonte_padrao, size=10, bold=True, color=COR_HEADER_TEXTO)
        cel_titulo.fill = PatternFill(start_color=COR_NEON, end_color=COR_NEON, fill_type="solid")
        cel_titulo.alignment = Alignment(horizontal="right", vertical="center")

        intervalo_matriz = (
            f"{get_column_letter(coluna_inicial)}{primeira_linha_dados}:"
            f"{get_column_letter(ultima_coluna)}{ultima_linha_dados}"
        )

        for col_idx in range(len(cnpjs)):
            letra_coluna = get_column_letter(coluna_inicial + col_idx)
            intervalo_coluna = f"{letra_coluna}{primeira_linha_dados}:{letra_coluna}{ultima_linha_dados}"

            cel = ws.cell(
                row=linha_risco,
                column=coluna_inicial + col_idx,
                value=f"=SUM({intervalo_coluna})/SUM({intervalo_matriz})",
            )
            cel.number_format = "0.00%"
            cel.font = Font(name=fonte_padrao, size=10, bold=True, color=COR_NEON)
            cel.alignment = Alignment(horizontal="center")

    ws.freeze_panes = ws.cell(row=primeira_linha_dados, column=coluna_inicial).coordinate
    _autofit_colunas(ws, largura_min=10, largura_max=22)


def gerar_excel_portfolio(
    fundos: list[dict],
    correlacao: dict | None = None,
    covariancia: dict | None = None,
    data_referencia: str | None = None,
) -> io.BytesIO:
    """
    Gera um arquivo .xlsx com a composição do portfólio (CNPJ, nome,
    rentabilidade, volatilidade e peso) e, em abas separadas, a matriz
    de correlação e a matriz de covariância entre os fundos selecionados.

    Parâmetros
    ----------
    fundos : list[dict]
        Cada item: { cnpj, nome, rentabilidade_36m, volatilidade_36m, peso }.
        `peso` é opcional (percentual, ex: 33.33 = 33,33%).
    correlacao : dict | None
        Matriz de correlação no formato { cnpj_coluna: { cnpj_linha: valor } },
        mesmo shape retornado por /api/portfolio/gerar.
    covariancia : dict | None
        Matriz de covariância, mesmo shape de `correlacao`: cada célula é
        correlacao(i,j) * volatilidade(i) * volatilidade(j) * peso(i) *
        peso(j) — o valor "real" (não normalizado) de cada par de fundos.
    data_referencia : str | None
        Data de referência usada no cálculo (apenas informativo no arquivo).

    Retorno
    -------
    io.BytesIO
        Buffer do arquivo .xlsx pronto para ser enviado via send_file.
    """
    wb = Workbook()

    _montar_aba_portfolio(wb, fundos, data_referencia)
    _montar_aba_correlacao(wb, fundos, correlacao or {})
    _montar_aba_covariancia(wb, fundos, covariancia or {})

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer